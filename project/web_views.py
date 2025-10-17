from django.contrib import messages
from django.shortcuts import render, redirect, HttpResponse
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.urls import reverse
from django.utils import timezone
from django.db.models import F, Sum
import calendar
from openpyxl import Workbook
from datetime import datetime
from django.db.models.functions import TruncDate
from django.db.models import Q
from project.forms.login_form import LoginForm
from project.forms.purchase_form import PurchaseForm
from project.forms.sale_form import SaleProductForm
from project.forms.user_form import UserForm
from project.forms.work_log_form import WorkLogForm, SalaryForm
from .models import *
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext_lazy as _

@login_required(login_url='website:login-view') 
def index(request):
    today = timezone.now().date()

    total_purchase = Purchase.objects.filter(create_date__date=today).aggregate(total=Sum('total_price'))['total'] or 0
    total_sales = SaleProduct.objects.filter(create_date__date=today).aggregate(total=Sum('total_price'))['total'] or 0
    users_count = User.objects.count()
    pending_issues = Purchase.objects.filter(create_date__date=today,pay_status=False).count()
    today_purchase = Purchase.objects.filter(create_date__date=today, pay_status=True).count()
    toal_price_pending_issues = Purchase.objects.filter(create_date__date=today,pay_status=False).aggregate(total=Sum('total_price'))['total'] or 0
    toal_price_today_purchase = Purchase.objects.filter(create_date__date=today, pay_status=True).aggregate(total=Sum('total_price'))['total'] or 0
    sale_pending_issues = SaleProduct.objects.filter(pay_status=False).count()
    today_sales = SaleProduct.objects.filter(create_date__date=today,  pay_status=True).count() or 0
    toal_price_pending_issues_sale = SaleProduct.objects.filter(create_date__date=today,pay_status=False).aggregate(total=Sum('total_price'))['total'] or 0
    toal_price_today_sale = SaleProduct.objects.filter(create_date__date=today, pay_status=True).aggregate(total=Sum('total_price'))['total'] or 0
    
    context = {
        'total_purchase': total_purchase,
        'total_sales': total_sales,
        'users_count': users_count,
        'pending_issues': pending_issues,
        'today_purchase': today_purchase,
        'toal_price_pending_issues': toal_price_pending_issues,
        'toal_price_today_purchase': toal_price_today_purchase,
        'sale_pending_issues': sale_pending_issues,
        'toal_price_pending_issues_sale': toal_price_pending_issues_sale,
        'toal_price_today_sale': toal_price_today_sale,
        'today_sales': today_sales,
    }
    return render(request, 'dashboard/index.html', context)

@login_required(login_url='website:login-view') 
def user_view(request):
    query = request.GET.get('q', '')
    users = User.objects.all()
    if query:
        users = users.filter(
            Q(username__icontains=query)
        )
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'user/index.html', {"users": page_obj})

@login_required(login_url='website:login-view') 
def user_create_view(request):
    if request.method == "POST":
        form = UserForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect("website:user-view")
    else:
        form = UserForm()
    return render(request, 'user/user.html', {'form': form,'title': _('Add New User')})

@login_required(login_url='website:login-view') 
def user_edit_view(request, pk):
    user = User.objects.filter(pk=pk).first()
    if request.method == "POST":
        form = UserForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            return redirect("website:user-view")
    else:
        initial_data = {}
        if user:
            initial_data['phone'] = user.profile.phone
            initial_data['profile'] = user.profile.profile
        form = UserForm(instance=user, initial=initial_data)
    return render(request, 'user/user.html', {'form': form,'title': _('Update user information')})

@login_required(login_url='website:login-view') 
def user_delete_view(request, pk):
    user = User.objects.get(pk=pk)
    if user:
        user.delete()
        return redirect("website:user-view")
    else:
        return redirect("website:user-view")


@login_required
def dashboard(request, pk):
    now = timezone.now()
    current_year = int(request.POST.get('year', now.year))
    current_month = int(request.POST.get('month', now.month))
    worklogs = WorkLog.objects.filter(user_id=pk, date__month=current_month,).order_by("-date")

    total_salary = sum(
        log.salary.amount for log in worklogs if hasattr(log, 'salary')
    )
    paginator = Paginator(worklogs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    months = list(range(1, 13))
    years = list(range(current_year - 3, current_year + 6))
    context={
                "logs": page_obj, 
                "total_salary": total_salary, 
                'months': months, 
                'years': years,
                "selected_month": current_month, 
                "selected_year": current_year
            }
    return render(request, "user/dashboard.html", context)

@login_required
def add_worklog(request, pk):
    if request.method == "POST":
        form = SalaryForm(request.POST)
        worklog_id = request.POST.get('worklog')  
        worklog = WorkLog.objects.filter(pk=worklog_id).first()
        print(worklog)
        if not worklog:
            messages.error(request, "WorkLog not found.")
            return redirect("website:user-view")
        if form.is_valid():
            if not worklog.is_leave:
                Salary.objects.create(worklog=worklog, amount=request.POST.get('amount'))
            else:
                Salary.objects.create(worklog=worklog, amount=0)
            worklog.paystatus = True
            worklog.save()
            return redirect(reverse("website:dashboard", args=[worklog.user_id]))
    else:
        form = SalaryForm()
    return render(request, "user/worklog_form.html", {"form": form, 'logid': pk})



@login_required(login_url='website:login-view') 
def purchase_view(request):
    query = request.GET.get('q', '')
    purchases = Purchase.objects.all().order_by('-id')
    if query:
        purchases = purchases.filter(
            Q(name__icontains=query) |
            Q(purchase_no__icontains=query)
        )
    paginator = Paginator(purchases, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'purchase/index.html', {'datas': page_obj})

@login_required(login_url='website:login-view') 
def purchase_create_view(request):
    if request.method == "POST":
        form = PurchaseForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            return redirect("website:purchase-view")
    else:
        form = PurchaseForm()
    return render(request, 'purchase/purchase.html', {'form': form, 'title': _('Add New Purchase Record')})

@login_required(login_url='website:login-view') 
def purchase_edit_view(request, pk):
    purchase = Purchase.objects.filter(pk=pk).first()
    if request.method == "POST":
        form = PurchaseForm(request.POST, instance=purchase, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('website:purchase-view') 
    else:
        form = PurchaseForm(instance=purchase, user=request.user)
    return render(request, 'purchase/purchase.html', {'form': form, 'title': _('Update Purchase Record'), 'action':  f"/purchase/{purchase.id}/edit"})

@login_required(login_url='website:login-view') 
def purchase_delete_view(request, pk):
    purchase = Purchase.objects.get(pk=pk)
    if purchase:
        RecoveryPurchase.objects.create(
            delete_by=request.user,
            user=purchase.user,
            name=purchase.name,
            purchase_no=purchase.purchase_no,
            amount=purchase.amount,
            price=purchase.price,
            total_price=purchase.total_price,
            pay_status=purchase.pay_status,
            type=purchase.type,
            create_date=purchase.create_date
        )
        purchase.delete()
        return redirect('website:purchase-view') 
    else:
        return redirect('website:purchase-view') 

@login_required(login_url='website:login-view') 
def sale_view(request):
    query = request.GET.get('q', '')
    sale = SaleProduct.objects.all().order_by('-id')
    if query:
        sale = sale.filter(
            Q(name__icontains=query) |
            Q(sale_no__icontains=query)
        )
    paginator = Paginator(sale, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'sale/index.html', {'datas': page_obj})

@login_required(login_url='website:login-view') 
def sale_create_view(request):
    if request.method == "POST":
        form = SaleProductForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            return redirect("website:sale-view")
    else:
        form = SaleProductForm()
    return render(request, 'sale/sale.html', {'form': form, 'title': _('Add New Sale Record')})

@login_required(login_url='website:login-view') 
def sale_edit_view(request, pk):
    sale = SaleProduct.objects.filter(pk=pk).first()
    if request.method == "POST":
        form = SaleProductForm(request.POST, instance=sale, user=request.user)
        if form.is_valid():
            form.save()
            return redirect('website:sale-view') 
    else:
        form = SaleProductForm(instance=sale, user=request.user)
    return render(request, 'sale/sale.html', {'form': form, 'title': _('Update Sale Record'), 'action':  f"/sale/{sale.id}/edit"})

@login_required(login_url='website:login-view') 
def sale_delete_view(request, pk):
    sale = SaleProduct.objects.get(pk=pk)
    if sale:
        RecoverySaleProduct.objects.create(
            delete_by=request.user,
            user=sale.user,
            name=sale.name,
            sale_no=sale.sale_no,
            amount=sale.amount,
            price=sale.price,
            total_price=sale.total_price,
            pay_status=sale.pay_status,
            type=sale.type,
            create_date=sale.create_date
        )
        sale.delete()
        return redirect('website:sale-view') 
    else:
        return redirect('website:sale-view') 

@login_required(login_url='website:login-view')    
def analysis_view(request):
    now = timezone.now()
    current_year = int(request.GET.get('year', now.year))
    current_month = int(request.GET.get('month', now.month))
    purchase_queryset = Purchase.objects.filter(
        create_date__year=current_year,
        create_date__month=current_month,
    )
    sale_queryset = SaleProduct.objects.filter(
        create_date__year=current_year,
        create_date__month=current_month,
    )
    purchase = (
        purchase_queryset
        .annotate(day=TruncDate('create_date'))
        .values('day')
        .annotate(total_price=Sum(F('amount') * F('price')))
        .order_by('day')
    )
    sale = (
        sale_queryset
        .annotate(day=TruncDate('create_date'))
        .values('day')
        .annotate(total_price=Sum(F('amount') * F('price')))
        .order_by('day')
    )
    purchase_dict = {item['day'].day: item['total_price'] for item in purchase}
    sale_dict = {item['day'].day: item['total_price'] for item in sale}
    num_days = calendar.monthrange(current_year, current_month)[1]
    purchase_list = []
    sale_list = []

    for day in range(1, num_days + 1):
        purchase_list.append({
            "day": str(day),
            "total_amount":int( purchase_dict.get(day, 0)) 
        })
    for day in range(1, num_days + 1):
        sale_list.append({
            "day": str(day),
            "total_amount":int( sale_dict.get(day, 0)) 
        })
    months = list(range(1, 13))
    years = list(range(current_year - 3, current_year + 6))
    context = {'purchases': purchase_list, 
               "selected_month": current_month, 
               "selected_year": current_year, 
               'months': months,
               'years': years,
               'sales': sale_list
               }
    return render(request, 'analysis/index.html', context)

@login_required(login_url='website:login-view') 
def report_view(request):
    date_str = request.POST.get('date')
    if date_str:
        date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        date = timezone.localdate()
    purchases = Purchase.objects.filter(create_date__date=date)
    sales = SaleProduct.objects.filter(create_date__date=date)
    
    purchase_paginator = Paginator(purchases, 10)
    purchase_page_number = request.GET.get('page')
    purchase_page_obj = purchase_paginator.get_page(purchase_page_number)

    sale_paginator = Paginator(sales, 10)
    sale_page_number = request.GET.get('page')
    sale_page_obj = sale_paginator.get_page(sale_page_number)
    context = {
        "today": date, 
        'sale_page_obj': sale_page_obj, 
        'sale_total_price': sales.aggregate(total=Sum(F('amount') * F('price')))['total'] or 0 ,
        'purchase_page_obj': purchase_page_obj, 
        'purchase_total_price': purchases.aggregate(total=Sum(F('amount') * F('price')))['total'] or 0 
        }
    return render(request, 'report/daily/index.html', context)

@login_required(login_url='website:login-view') 
def download_purchase_daily_csv(request):
    date_str = request.GET.get('date', timezone.localdate().isoformat())
    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    purchases = Purchase.objects.filter(create_date__date=date_obj)
    
    columns = ['ID', 'Product No', 'Product', 'Amount', 'Price', 'Total Price', 'Created At']
    
    return generate_excel_response(purchases, 'purchases', columns)

@login_required(login_url='website:login-view') 
def download_sale_daily_csv(request):
    date_str = request.GET.get('date', timezone.localdate().isoformat())
    date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
    sales = SaleProduct.objects.filter(create_date__date=date_obj)
    
    columns = ['ID', 'Product No', 'Product', 'Amount', 'Price', 'Total Price', 'Created At']
    
    return generate_excel_response(sales, 'sales', columns)

@login_required(login_url='website:login-view') 
def download_purchase_monthly_csv(request):
    now = timezone.now()
    current_year = int(request.POST.get('year', now.year))
    current_month = int(request.POST.get('month', now.month))
    purchases = Purchase.objects.filter(
        create_date__year=current_year,
        create_date__month=current_month,
    )
    
    columns = ['ID', 'Product No', 'Product', 'Amount', 'Price', 'Total Price', 'Created At']
    
    return generate_excel_response(purchases, 'purchases', columns)

@login_required(login_url='website:login-view') 
def download_sale_monthly_csv(request):
    now = timezone.now()
    current_year = int(request.POST.get('year', now.year))
    current_month = int(request.POST.get('month', now.month))
    sales = SaleProduct.objects.filter(
        create_date__year=current_year,
        create_date__month=current_month,
    )
    
    columns = ['ID', 'Product No', 'Product', 'Amount', 'Price', 'Total Price', 'Created At']
    
    return generate_excel_response(sales, 'sales', columns)

@login_required(login_url='website:login-view') 
def report_monthly_view(request):
    now = timezone.now()
    current_year = int(request.POST.get('year', now.year))
    current_month = int(request.POST.get('month', now.month))
    purchase_queryset = Purchase.objects.filter(
        create_date__year=current_year,
        create_date__month=current_month,
    )
    sale_queryset = SaleProduct.objects.filter(
        create_date__year=current_year,
        create_date__month=current_month,
    )
    purchase_paginator = Paginator(purchase_queryset, 10)
    purchase_page_number = request.GET.get('page')
    purchase_page_obj = purchase_paginator.get_page(purchase_page_number)

    sale_paginator = Paginator(sale_queryset, 10)
    sale_page_number = request.GET.get('page')
    sale_page_obj = sale_paginator.get_page(sale_page_number)

    months = list(range(1, 13))
    years = list(range(current_year - 3, current_year + 6))
    context = {'purchases': purchase_page_obj, 
               'purchase_total_price':  purchase_queryset.aggregate(total=Sum('total_price'))['total'] or 0 ,
               "selected_month": current_month, 
               "selected_year": current_year, 
               'months': months,
               'years': years,
               'sales': sale_page_obj,
               'sale_total_price':  sale_queryset.aggregate(total=Sum('total_price'))['total'] or 0 ,
               }
    return render(request, 'report/monthly/index.html', context)

def logout_view(request):
    logout(request)
    return redirect('website:index')


def login_view(request):
    if request.method =="POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(username=request.POST.get('username'), password=request.POST.get('password'))
            if user and user.is_superuser:
                login(request, user)
                return redirect('website:index')
            else:
                messages.error(request, "Invalid email/user account or password")
    else:
        form=LoginForm()
    return render(request, 'auth/login.html', {'form': form})





from openpyxl.styles import Font, Alignment, PatternFill
def generate_excel_response(queryset, filename_prefix, columns):
    """
    Reusable Excel export function for Purchases or Sales.
    Includes a Total row at the bottom.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"{filename_prefix.capitalize()} {timezone.localdate()}"
    
    # Write header
    ws.append(columns)

    # Dynamic field name (purchase_no or sale_no)
    for_no = 'purchase_no' if filename_prefix == 'purchases' else 'sale_no'

    total_sum = 0
    total_amount = 0

    # Data rows
    for obj in queryset:
        total_amount += obj.amount
        total_sum += obj.total_price
        ws.append([
            obj.id,
            getattr(obj, for_no,''),
            getattr(obj, 'name', 'N/A'),
            obj.amount,
            obj.price,
            obj.total_price,
            obj.create_date.strftime('%Y-%m-%d'),
        ])

    ws.append([])

    last_row = ws.max_row + 1
    ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row + 1, end_column=3)
    ws.merge_cells(start_row=last_row, start_column=4, end_row=last_row + 1, end_column=4)
    ws.merge_cells(start_row=last_row, start_column=6, end_row=last_row + 1, end_column=6)

    # Set values
    ws.cell(row=last_row, column=1, value="TOTAL SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=last_row, column=4, value=total_amount).font = Font(bold=True, size=12)
    ws.cell(row=last_row, column=6, value=total_sum).font = Font(bold=True, size=12)

    # Alignment
    ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=last_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=last_row, column=6).alignment = Alignment(horizontal="center", vertical="center")

    # Background fill (warning yellow)
    fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    for row in range(last_row, last_row + 2):
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = fill


    ws.cell(row=last_row, column=1).alignment = Alignment(horizontal="center")

    # Response
    date_str = timezone.localdate().isoformat()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{date_str}.xlsx"'
    wb.save(response)
    return response
